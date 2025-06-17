import openpyxl
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from django.views import View
from django.contrib.auth.models import User
from openpyxl.styles import Font, PatternFill
from django.contrib.auth.models import User, Group
from .models import ContactDetails, Devise, DeviseApis, APICountThreshold, ColumnName, DeviseLocation, DeviseApisFields, SOIL_LIFE_FIELDS, ATMO_SENSE_FIELDS, SOIL_SAATHI_FIELDS
from datetime import datetime

DEVICE_FIELDS_MAP = {
    'soilsaathi': SOIL_SAATHI_FIELDS,
    'atmo_sense': ATMO_SENSE_FIELDS,
    'soil_life': SOIL_LIFE_FIELDS,
}


class ExportUsersView(View):
    """Exports only users in the 'deviseowner' group with formatting."""
    
    def get(self, request, *args, **kwargs):
        try:
            group = Group.objects.get(name="deviseowner")
            users = User.objects.filter(groups=group)  # Filter only users in this group
        except Group.DoesNotExist:
            users = User.objects.none()  # If group does not exist, return an empty sheet

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deviseowner Users"

        # Define header style
        bold_font = Font(bold=True, color="FFFFFF")  # White text
        fill_color = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background

        # Add headers
        headers = ["Username", "First Name", "Last Name", "Email"]
        ws.append(headers)

        # Apply formatting to header row
        for col_num, header in enumerate(headers, 1):
            cell      = ws.cell(row=1, column=col_num)
            cell.font = bold_font
            cell.fill = fill_color

        # Add user data (only for 'deviseowner' group)
        for user in users:
            ws.append([user.username, user.first_name, user.last_name, user.email])

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter  # Get column letter (A, B, C, etc.)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="deviseowner_users.xlsx"'
        wb.save(response)
        return response

class ExportUsersAndDevicesView(View):
    """Exports users in one sheet and their devices in another sheet (hierarchical format)."""
    
    def get(self, request, username=None, *args, **kwargs):
        wb = openpyxl.Workbook()

        # ============ USERS SHEET ============
        ws_users = wb.active
        ws_users.title = "Users"

        # Header styles
        bold_font = Font(bold=True, color="FFFFFF")
        fill_color = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Define user headers
        user_headers = ["Username", "First Name", "Last Name", "Email"]
        ws_users.append(user_headers)

        # Apply header formatting
        for col_num, header in enumerate(user_headers, 1):
            cell = ws_users.cell(row=1, column=col_num)
            cell.font = bold_font
            cell.fill = fill_color

        if username:
            users = User.objects.filter(username=username)
        else:
            group = Group.objects.get(name="deviseowner")
            users = User.objects.filter(groups=group)

        for user in users:
            ws_users.append([user.username, user.first_name, user.last_name, user.email])

        # Auto-adjust column width
        for col in ws_users.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws_users.column_dimensions[col[0].column_letter].width = max_length + 2

        # ============ DEVICES SHEET ============
        if not username:
            ws_devices = wb.create_sheet(title="Devices")

            # Device headers (User info + Device info)
            device_headers = [
                "Username", "First Name", "Last Name", "Email",
                "Device Name", "Serial No", "Device ID", "Chipset No", "Email", "Phone",
                "Address1", "Address2", "Purchase Date", "Time of Sale", "Warranty",
                "Amount Paid", "Balance Amount", "Land", "Created At", "Device Type"
            ]
            ws_devices.append(device_headers)

            # Apply formatting to header row
            for col_num, header in enumerate(device_headers, 1):
                cell      = ws_devices.cell(row=1, column=col_num)
                cell.font = bold_font
                cell.fill = fill_color

            for user in users:
                ws_devices.append([user.username, user.first_name, user.last_name, user.email] + [""] * (len(device_headers) - 4))
                user_devices = Devise.objects.filter(user=user)

                for device in user_devices:
                    row_data = [
                        "", "", "", "",
                        device.name, device.serial_no, device.devise_id, device.chipset_no, device.email,
                        device.phone, device.address1, device.address2,
                        device.purchase_date.replace(tzinfo=None) if isinstance(device.purchase_date, datetime) else device.purchase_date,
                        device.time_of_sale.replace(tzinfo=None) if isinstance(device.time_of_sale, datetime) else device.time_of_sale,
                        device.warrenty.replace(tzinfo=None) if isinstance(device.warrenty, datetime) else device.warrenty,
                        device.amount_paid, device.balance_amount, device.land,
                        device.created_at.replace(tzinfo=None) if isinstance(device.created_at, datetime) else device.created_at,
                        device.devise_type
                    ]
                    ws_devices.append(row_data)

            # Auto-adjust column width
            for col in ws_devices.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws_devices.column_dimensions[col[0].column_letter].width = max_length + 2
        else:

            # ============ USER DEVICE DETAILS SHEET ============
            ws_user_devices = wb.create_sheet(title="User Device Details")

            # Headers for User Device Details
            user_device_headers = [
                "Username", "Device Name", "Serial No", "Device ID", "Chipset No", "Email", "Phone",
                "Address1", "Address2", "Purchase Date", "Time of Sale", "Warranty",
                "Amount Paid", "Balance Amount", "Land", "Created At", "Device Type"
            ]
            ws_user_devices.append(user_device_headers)

            # Apply formatting to header row
            for col_num, header in enumerate(user_device_headers, 1):
                cell = ws_user_devices.cell(row=1, column=col_num)
                cell.font = bold_font
                cell.fill = fill_color

            for user in users:
                user_devices = Devise.objects.filter(user=user)
                for device in user_devices:
                    row_data = [
                        user.username, device.name, device.serial_no, device.devise_id, device.chipset_no, device.email,
                        device.phone, device.address1, device.address2,
                        device.purchase_date.replace(tzinfo=None) if isinstance(device.purchase_date, datetime) else device.purchase_date,
                        device.time_of_sale.replace(tzinfo=None) if isinstance(device.time_of_sale, datetime) else device.time_of_sale,
                        device.warrenty.replace(tzinfo=None) if isinstance(device.warrenty, datetime) else device.warrenty,
                        device.amount_paid, device.balance_amount, device.land,
                        device.created_at.replace(tzinfo=None) if isinstance(device.created_at, datetime) else device.created_at,
                        device.devise_type
                    ]
                    ws_user_devices.append(row_data)

            # Auto-adjust column width
            for col in ws_user_devices.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws_user_devices.column_dimensions[col[0].column_letter].width = max_length + 2

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="users_devices_details.xlsx"'
        wb.save(response)
        return response
        
class ExportDeviceApisView(View):
    """Exports all APIs of a specific device along with device details and user details."""

    def get(self, request, device_id, *args, **kwargs):
        try:
            device = Devise.objects.get(id=device_id)
            user = device.user  # Fetch user details
            fields = DEVICE_FIELDS_MAP.get(device.devise_type, {})

            # Select the correct model
            if device.devise_type == 'soilsaathi':
                apis = DeviseApis.objects.filter(device=device)
            else:
                apis = DeviseApisFields.objects.filter(device=device)

            total_api_calls = apis.count()
        except Devise.DoesNotExist:
            return HttpResponse("Device not found", status=404)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Device APIs"

        # Define header style
        bold_font = Font(bold=True, color="FFFFFF")
        fill_color = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # ============ DEVICE & USER DETAILS ============
        device_details = [
            ["Device Name", device.name],
            ["Device ID", device.devise_id],
            ["Serial No", device.serial_no],
            ["Chipset No", device.chipset_no],
            ["Device Type", device.devise_type],
            ["User", user.username],
            ["Email", user.email],
            ["Phone", device.phone],
            ["Total API Calls", total_api_calls]
        ]

        for row in device_details:
            ws.append(row)

        ws.append([])  # Blank row for spacing

        # ============ API HEADERS ============
        headers = list(fields.values())
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.font = bold_font
            cell.fill = fill_color

        # ============ API DATA ============
        for api in apis:
            row_data = []
            for field in fields.keys():
                value = getattr(api, field, "N/A")

                # Ensure datetime fields are timezone-naïve
                if isinstance(value, datetime):
                    value = value.replace(tzinfo=None)

                row_data.append(value)

            ws.append(row_data)

        # Auto-adjust column width
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="device_{device_id}_apis.xlsx"'
        wb.save(response)
        return response